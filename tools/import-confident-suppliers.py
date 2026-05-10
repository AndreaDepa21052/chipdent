#!/usr/bin/env python3
"""Regenera src/Chipdent.Web/Infrastructure/Mongo/ConfidentImportSeedData.cs a
partire da FileRaw/FORNITORI Confident.xlsx.

Uso (dalla root del repo):
    python3 tools/import-confident-suppliers.py

Sheet sorgenti:
    - "fornitori": anagrafica aziende (codice F####)
    - "medici":    anagrafica dottori   (codice D####)

Per ogni dottore importato il seeder runtime crea anche un Fornitore-ombra
collegato (vedi FornitoreOmbraService), così l'anagrafica fornitori comprende
sia le aziende che i dottori.
"""
from __future__ import annotations

import re
from pathlib import Path

import openpyxl  # type: ignore[import-not-found]


REPO_ROOT = Path(__file__).resolve().parents[1]
XLSX_PATH = REPO_ROOT / "FileRaw" / "FORNITORI Confident.xlsx"
OUT_PATH = REPO_ROOT / "src" / "Chipdent.Web" / "Infrastructure" / "Mongo" / "ConfidentImportSeedData.cs"


def cs_str(value, default: str = "null") -> str:
    if value is None:
        return default
    s = str(value).strip().replace("\xa0", " ").strip()
    if not s:
        return default
    s = s.replace("\\", "\\\\").replace("\"", "\\\"")
    return f'"{s}"'


def split_emails(*values) -> str | None:
    out: list[str] = []
    for v in values:
        if v is None:
            continue
        s = str(v).strip().replace("\xa0", " ").strip()
        if not s:
            continue
        for piece in re.split(r"[;,]", s):
            piece = piece.strip()
            if piece and "@" in piece and piece not in out:
                out.append(piece)
    return "; ".join(out) if out else None


def build_fornitori_lines(ws) -> list[str]:
    lines: list[str] = []
    seq = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        nome = row[0]
        if nome is None or not str(nome).strip():
            continue
        seq += 1
        codice = f"F{seq:04d}"
        nome = str(nome).strip().replace("\xa0", " ")
        email = split_emails(row[1], row[2])
        note_parts: list[str] = []
        if row[5] and str(row[5]).strip():
            note_parts.append(str(row[5]).strip().replace("\xa0", " "))
        if row[7] and str(row[7]).strip():
            note_parts.append(str(row[7]).strip().replace("\xa0", " "))
        if row[4] and str(row[4]).strip():
            note_parts.append(f"Mail sbagliata: {str(row[4]).strip()}")
        note = " — ".join(note_parts) if note_parts else None
        lines.append(
            f'        new Fornitore {{ TenantId = tenantId, Codice = "{codice}", '
            f'RagioneSociale = {cs_str(nome, "string.Empty")}, '
            f'EmailContatto = {cs_str(email)}, '
            f'Note = {cs_str(note)}, '
            f'CategoriaDefault = CategoriaSpesa.AltreSpeseFisse, '
            f'Stato = StatoFornitore.Attivo, '
            f'TerminiPagamentoGiorni = 30, '
            f'BasePagamento = BasePagamento.DataFattura }},'
        )
    return lines


def build_dottori_lines(ws) -> list[str]:
    lines: list[str] = []
    seq = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        nome = row[0]
        if nome is None or not str(nome).strip():
            continue
        seq += 1
        codice = f"D{seq:04d}"
        nome = str(nome).strip().replace("\xa0", " ")
        parts = nome.split()
        if len(parts) >= 2:
            cognome = parts[0]
            nomep = " ".join(parts[1:])
        else:
            cognome = nome
            nomep = ""
        email = split_emails(row[1])
        lines.append(
            f'        new Dottore {{ TenantId = tenantId, Codice = "{codice}", '
            f'Nome = {cs_str(nomep, "string.Empty")}, Cognome = {cs_str(cognome, "string.Empty")}, '
            f'Email = {cs_str(email, "string.Empty")}, '
            f'Specializzazione = string.Empty, NumeroAlbo = string.Empty, '
            f'TipoContratto = TipoContratto.Collaborazione, '
            f'DataAssunzione = seedDate, '
            f'Attivo = true }},'
        )
    return lines


def main() -> None:
    wb = openpyxl.load_workbook(str(XLSX_PATH), data_only=True)
    fornitori_lines = build_fornitori_lines(wb["fornitori"])
    dottori_lines = build_dottori_lines(wb["medici"])

    header = (
        "// AUTO-GENERATED from FileRaw/FORNITORI Confident.xlsx\n"
        "// Do NOT edit by hand — rigenerare con tools/import-confident-suppliers.py\n"
        "using Chipdent.Web.Domain.Entities;\n\n"
        "namespace Chipdent.Web.Infrastructure.Mongo;\n\n"
        "internal static class ConfidentImportSeedData\n"
        "{\n"
        "    /// <summary>Lista fornitori (aziende) importati dal foglio \"fornitori\" del file Confident.</summary>\n"
        "    public static IReadOnlyList<Fornitore> BuildFornitori(string tenantId) => new[]\n"
        "    {\n"
    )
    mid = (
        "    };\n\n"
        "    /// <summary>Lista dottori importati dal foglio \"medici\" del file Confident.\n"
        "    /// Ognuno avrà un Fornitore-ombra agganciato (vedi <see cref=\"Chipdent.Web.Infrastructure.Sepa.FornitoreOmbraService\"/>).</summary>\n"
        "    public static IReadOnlyList<Dottore> BuildDottori(string tenantId)\n"
        "    {\n"
        "        var seedDate = new DateTime(2020, 1, 1, 0, 0, 0, DateTimeKind.Utc);\n"
        "        return new[]\n"
        "        {\n"
    )
    footer = "        };\n    }\n}\n"

    with OUT_PATH.open("w", encoding="utf-8") as f:
        f.write(header)
        for line in fornitori_lines:
            f.write(line + "\n")
        f.write(mid)
        for line in dottori_lines:
            f.write(line + "\n")
        f.write(footer)

    print(f"Wrote {OUT_PATH.relative_to(REPO_ROOT)}: "
          f"{len(fornitori_lines)} fornitori, {len(dottori_lines)} dottori")


if __name__ == "__main__":
    main()
